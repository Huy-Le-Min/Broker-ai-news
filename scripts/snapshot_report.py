# -*- coding: utf-8 -*-
"""
Snapshot report — tạo Excel từ dữ liệu do CLAUDE fetch (WebSearch/WebFetch).
Dùng cho keyword mà Python local KHÔNG vào được nguồn (vàng, CPI tháng, dầu...).

Cách dùng:  py scripts/snapshot_report.py path/to/data.json
JSON schema:
{
  "key": "VANG",
  "title": "GIÁ VÀNG SJC",
  "source": "...",
  "as_of": "10/06/2026",
  "table": [["Chỉ tiêu","Giá trị"], ["...","..."]],
  "bullets": ["...", "..."]
}
"""
import sys, os, io, json, contextlib
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "output")
os.makedirs(OUT_DIR, exist_ok=True)


def write(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Tổng quan"
    ws["A1"] = data.get("title", data.get("key", "BÁO CÁO"))
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
    ws["A2"] = f"Nguồn: {data.get('source','n/a')} · Số liệu tới: {data.get('as_of','n/a')} · Xuất: {datetime.today():%d/%m/%Y}"
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    row = 4
    table = data.get("table") or []
    if table:
        ws.cell(row=row, column=1, value="SỐ LIỆU").font = Font(bold=True, size=12, color="1F4E78")
        row += 1
        blue = PatternFill("solid", fgColor="DDEBF7")
        for i, tr in enumerate(table):
            for j, val in enumerate(tr):
                c = ws.cell(row=row, column=1 + j, value=val)
                if i == 0:
                    c.font = Font(bold=True); c.fill = blue
            row += 1
        row += 1

    bullets = data.get("bullets") or []
    if bullets:
        ws.cell(row=row, column=1, value="INSIGHTS").font = Font(bold=True, size=12, color="1F4E78")
        row += 1
        for b in bullets:
            cell = ws.cell(row=row, column=1, value="• " + b)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.row_dimensions[row].height = 30
            row += 1
        row += 1

    ws.cell(row=row, column=1, value="Disclaimer: Báo cáo tự động, mang tính tham khảo, không phải khuyến nghị đầu tư. "
            "Số liệu từ nguồn công khai do Claude tổng hợp tại thời điểm cập nhật.").font = Font(italic=True, size=8, color="808080")
    ws.column_dimensions["A"].width = 34
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 18

    out_path = os.path.join(OUT_DIR, f"{data.get('key','snapshot')}_{datetime.today():%Y%m%d}.xlsx")
    wb.save(out_path)
    return out_path


def main():
    path = sys.argv[1]
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    out = write(data)
    print("Đã xuất:", out)


if __name__ == "__main__":
    main()
