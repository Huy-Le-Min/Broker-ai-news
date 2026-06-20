# -*- coding: utf-8 -*-
"""
Automation 1 — Keyword(s) -> 1 Excel workbook nhiều sheet + insight.
Cách dùng:
    py scripts/automation1_keyword.py HPG
    py scripts/automation1_keyword.py CPI GDP "ty gia" HPG VCB
Tầng A (tự động, Python local): cổ phiếu (vnstock) + vĩ mô THEO NĂM (World Bank).
Tầng B (vàng/dầu/CPI tháng): dùng scripts/snapshot_report.py với data Claude fetch.
"""
import sys, os, io, re, json, time, contextlib, unicodedata
from datetime import datetime, timedelta
import requests

os.environ.setdefault("PYTHONUTF8", "1")
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
    import pandas as pd
    from vnstock import Trading
    from vnstock.api.quote import Quote

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "output")
DATA_DIR = os.path.join(ROOT, "data")
os.makedirs(OUT_DIR, exist_ok=True)

FULL = False  # --full: kèm dữ liệu ngày chi tiết


def load_macro_latest():
    """Số vĩ mô mới nhất (Tầng B) do Claude fetch, bổ sung cho World Bank (theo năm, trễ)."""
    p = os.path.join(DATA_DIR, "macro_latest.json")
    if os.path.exists(p):
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


MACRO_LATEST = load_macro_latest()


def load_stock_valuation():
    """Định giá cổ phiếu (P/E, P/B, ROE, P/E ngành) do Claude fetch (Tầng B). Tùy chọn."""
    p = os.path.join(DATA_DIR, "stock_valuation.json")
    if os.path.exists(p):
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


STOCK_VAL = load_stock_valuation()

# Cổ phiếu -> chỉ số ngành HOSE để so sánh
SECTOR_INDEX = {
    "Ngân hàng": ("VNFIN", "Tài chính (VNFIN)"),
    "Tài chính": ("VNFIN", "Tài chính (VNFIN)"),
    "Chứng khoán": ("VNFIN", "Tài chính (VNFIN)"),
    "Bảo hiểm": ("VNFIN", "Tài chính (VNFIN)"),
    "Bất động sản": ("VNREAL", "Bất động sản (VNREAL)"),
    "Thép": ("VNMAT", "Nguyên vật liệu (VNMAT)"),
    "Nguyên vật liệu": ("VNMAT", "Nguyên vật liệu (VNMAT)"),
    "Bán lẻ": ("VNCONS", "Tiêu dùng (VNCONS)"),
    "Tiêu dùng": ("VNCONS", "Tiêu dùng (VNCONS)"),
}
TICKER_SECTOR = {
    "VCB": "Ngân hàng", "ACB": "Ngân hàng", "BID": "Ngân hàng", "CTG": "Ngân hàng",
    "TCB": "Ngân hàng", "MBB": "Ngân hàng", "VPB": "Ngân hàng",
    "VIC": "Bất động sản", "VHM": "Bất động sản", "NVL": "Bất động sản", "DXG": "Bất động sản", "KDH": "Bất động sản",
    "HPG": "Thép", "HSG": "Thép", "NKG": "Thép",
    "MWG": "Bán lẻ", "FRT": "Bán lẻ", "PNJ": "Bán lẻ",
}


def load_portfolio_sectors():
    p = os.path.join(ROOT, "clients", "_me", "portfolio.csv")
    out = {}
    if os.path.exists(p):
        try:
            d = pd.read_csv(p)
            for _, r in d.iterrows():
                if pd.notna(r.get("ticker")) and pd.notna(r.get("nganh")):
                    out[str(r["ticker"]).upper()] = str(r["nganh"]).strip()
        except Exception:
            pass
    return out


PORTFOLIO_SECTORS = load_portfolio_sectors()


# ----------------- tiện ích -----------------
def pct(a, b):
    if b is None or b == 0 or pd.isna(b) or pd.isna(a):
        return None
    return (a / b - 1) * 100


def fmt_pct(x):
    return "n/a" if x is None else f"{x:+.1f}%"


def fmt_vnd(x):
    return "n/a" if pd.isna(x) else f"{x:,.0f} đ"


# ----------------- CỔ PHIẾU (vnstock) -----------------
def _hist(symbol, years=6):
    end = datetime.today()
    start = end - timedelta(days=365 * years)
    last_err = None
    for attempt in range(4):  # retry + backoff tránh rate-limit vnstock
        try:
            with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
                df = Quote(symbol=symbol, source="VCI").history(
                    start=start.strftime("%Y-%m-%d"), end=end.strftime("%Y-%m-%d"), interval="1D")
            df["time"] = pd.to_datetime(df["time"])
            return df.set_index("time").sort_index()
        except Exception as e:
            last_err = e
            time.sleep(2.0 * (attempt + 1))
    raise last_err


_INDEX_CACHE = {}


def get_stock_daily(ticker, years=6):
    df = _hist(ticker, years)
    for c in ["open", "high", "low", "close"]:
        df[c] = df[c] * 1000.0
    return df


def get_index_close(symbol):
    if symbol in _INDEX_CACHE:
        return _INDEX_CACHE[symbol]
    try:
        s = _hist(symbol)["close"]
    except Exception:
        s = None
    _INDEX_CACHE[symbol] = s
    return s


_LISTED_CACHE = {}


def prefetch_listed_shares(tickers):
    """Gọi price_board 1 lần cho nhiều mã (tránh nhiều request)."""
    todo = [t for t in tickers if t not in _LISTED_CACHE]
    if not todo:
        return
    try:
        with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
            tb = Trading(source="VCI").price_board(todo)
        for _, r in tb.iterrows():
            L = r["listing"]
            sym = str(L["symbol"]).upper()
            _LISTED_CACHE[sym] = float(L["listed_share"]) if pd.notna(L.get("listed_share")) else None
    except Exception:
        pass
    for t in todo:
        _LISTED_CACHE.setdefault(t, None)


def get_listed_share(tk):
    if tk not in _LISTED_CACHE:
        prefetch_listed_shares([tk])
    return _LISTED_CACHE.get(tk)


def period_returns(close):
    """Lợi suất 1 tháng (~21 phiên), YTD, 1 năm (~252 phiên)."""
    if close is None or len(close) == 0:
        return {}
    last = close.iloc[-1]
    ytd_base = close[close.index.year == close.index[-1].year].iloc[0]

    def r(n):
        return pct(last, close.iloc[-1 - n]) if len(close) > n else None
    return {"1 tháng": r(21), "YTD": pct(last, ytd_base), "1 năm": r(252)}


def build_periodic(df):
    m = pd.DataFrame({
        "Giá đóng cửa": df["close"].resample("ME").last(),
        "Cao nhất": df["high"].resample("ME").max(),
        "Thấp nhất": df["low"].resample("ME").min(),
        "KLGD bình quân": df["volume"].resample("ME").mean(),
    })
    m["MoM"] = m["Giá đóng cửa"].pct_change() * 100
    m["YoY"] = m["Giá đóng cửa"].pct_change(12) * 100
    m = m.tail(12)
    q = pd.DataFrame({
        "Giá đóng cửa": df["close"].resample("QE").last(),
        "Cao nhất": df["high"].resample("QE").max(),
        "Thấp nhất": df["low"].resample("QE").min(),
        "KLGD bình quân": df["volume"].resample("QE").mean(),
    })
    q["QoQ"] = q["Giá đóng cửa"].pct_change() * 100
    q["YoY"] = q["Giá đóng cửa"].pct_change(4) * 100
    q = q.tail(8)
    y = pd.DataFrame({
        "Giá đóng cửa": df["close"].resample("YE").last(),
        "Cao nhất": df["high"].resample("YE").max(),
        "Thấp nhất": df["low"].resample("YE").min(),
        "KLGD bình quân": df["volume"].resample("YE").mean(),
    })
    y["YoY"] = y["Giá đóng cửa"].pct_change() * 100
    y = y.tail(5)
    return m, q, y


def build_insights(ticker, df):
    close = df["close"]
    last = close.iloc[-1]
    last_date = close.index[-1].strftime("%d/%m/%Y")

    def ago(n):
        return close.iloc[-1 - n] if len(close) > n else None

    w52 = close.tail(252)
    hi, lo = w52.max(), w52.min()
    ytd_base = close[close.index.year == close.index[-1].year].iloc[0]
    vol20 = df["volume"].tail(20).mean()
    vol1y = df["volume"].tail(252).mean()
    ma50 = close.tail(50).mean()
    ma200 = close.tail(200).mean()
    trend = "tăng (giá > MA50 > MA200)" if last > ma50 > ma200 else (
        "giảm (giá < MA50 < MA200)" if last < ma50 < ma200 else "đi ngang/chưa rõ xu hướng")
    return [
        f"Giá đóng cửa mới nhất ({last_date}): {fmt_vnd(last)}.",
        f"Thay đổi: 1 phiên {fmt_pct(pct(last, ago(1)))}, ~1 tuần {fmt_pct(pct(last, ago(5)))}, "
        f"~1 tháng {fmt_pct(pct(last, ago(21)))}, YTD {fmt_pct(pct(last, ytd_base))}, "
        f"~1 năm {fmt_pct(pct(last, ago(252)))}.",
        f"Vùng 52 tuần: {fmt_vnd(lo)} – {fmt_vnd(hi)}; cách đỉnh 52T {fmt_pct(pct(last, hi))}.",
        f"Thanh khoản: KLGD BQ 20 phiên {vol20:,.0f} vs BQ 1 năm {vol1y:,.0f} "
        f"({fmt_pct(pct(vol20, vol1y))}).",
        f"Xu hướng kỹ thuật: {trend} (MA50 {fmt_vnd(ma50)}, MA200 {fmt_vnd(ma200)}).",
    ]


def build_stock_section(tk):
    df = get_stock_daily(tk)
    m, q, y = build_periodic(df)
    bullets = build_insights(tk, df)
    last_close = df["close"].iloc[-1]

    # Vốn hóa = giá × số CP niêm yết
    ls = get_listed_share(tk)
    if ls:
        bullets.append(f"Vốn hóa: ~{ls * last_close / 1e9:,.0f} tỷ đồng "
                       f"({ls / 1e9:,.2f} tỷ CP niêm yết × giá).")

    # Định giá (P/E, P/B, ROE...) — Tầng B nếu có
    val = STOCK_VAL.get(tk)
    if val:
        if val.get("valuation"):
            txt = val["valuation"]
        else:
            parts = [f"{lbl} {val[k]}" for k, lbl in
                     [("pe", "P/E"), ("pb", "P/B"), ("roe", "ROE"), ("eps", "EPS")] if val.get(k) is not None]
            if val.get("industry_pe") is not None:
                parts.append(f"P/E ngành {val['industry_pe']}")
            txt = ", ".join(parts)
        if txt:
            bullets.append(f"Định giá (ước tính): {txt} "
                           f"(Nguồn: {val.get('source', '')}, {val.get('as_of', '')}).")

    # So sánh với chỉ số ngành + VN-Index
    nganh = PORTFOLIO_SECTORS.get(tk) or TICKER_SECTOR.get(tk)
    sect = SECTOR_INDEX.get(nganh) if nganh else None
    st_ret = period_returns(df["close"])
    sect_ret = period_returns(get_index_close(sect[0])) if sect else {}
    vni_ret = period_returns(get_index_close("VNINDEX"))
    sect_label = sect[1] if sect else "Ngành (n/a)"
    cmp_df = pd.DataFrame([[k, st_ret.get(k), sect_ret.get(k), vni_ret.get(k)]
                           for k in ["1 tháng", "YTD", "1 năm"]],
                          columns=["Kỳ", tk, sect_label, "VN-Index"])
    if sect and st_ret.get("YTD") is not None and sect_ret.get("YTD") is not None:
        rel = "MẠNH hơn" if st_ret["YTD"] > sect_ret["YTD"] else "YẾU hơn"
        bullets.append(f"So ngành (YTD): {tk} {fmt_pct(st_ret['YTD'])} vs {sect_label} "
                       f"{fmt_pct(sect_ret['YTD'])} → {rel} ngành; VN-Index {fmt_pct(vni_ret.get('YTD'))}.")

    # 1 sheet/mã: các block xếp chồng cho gọn (bỏ dữ liệu ngày trừ khi --full)
    blocks = [
        ("So với ngành (%)", cmp_df, "compare"),
        ("Theo năm (5Y)", y.reset_index().rename(columns={"time": "Năm"}), "stockperiod"),
        ("Theo quý (2Y)", q.reset_index().rename(columns={"time": "Quý"}), "stockperiod"),
        ("Theo tháng (12M)", m.reset_index().rename(columns={"time": "Tháng"}), "stockperiod"),
    ]
    if FULL:
        blocks.append(("Dữ liệu ngày (1Y)", df.reset_index().tail(260), "stockraw"))
    return {"key": tk, "title": f"CỔ PHIẾU {tk}",
            "source": "vnstock 4.0.4 · dữ liệu sàn VCI (Vietcap)",
            "source_url": "https://vnstocks.com",
            "bullets": bullets, "sheet": (tk, blocks),
            "note": f"{len(df)} phiên {df.index[0]:%d/%m/%Y}–{df.index[-1]:%d/%m/%Y}"}


def build_latest_only_section(group):
    """Chỉ tiêu vĩ mô chỉ có số mới nhất (IIP, bán lẻ, FDI, XNK) — không có chuỗi World Bank."""
    latest = MACRO_LATEST.get(group)
    if not latest:
        return None
    return {"key": group, "title": f"VĨ MÔ {group}",
            "source": latest.get("source", ""), "source_url": latest.get("url", ""),
            "bullets": [f"MỚI NHẤT ({latest.get('as_of', '')}) — {latest['latest']}."],
            "sheet": None, "note": ""}


# ----------------- VĨ MÔ (World Bank) -----------------
WB_BASE = "https://api.worldbank.org/v2/country/VNM/indicator/{ind}"
MACRO_ALIASES = {
    "CPI": "CPI", "LAM PHAT": "CPI", "INFLATION": "CPI",
    "GDP": "GDP", "TANG TRUONG": "GDP", "TANG TRUONG GDP": "GDP",
    "TY GIA": "FX", "TYGIA": "FX", "USD": "FX", "USDVND": "FX", "USD/VND": "FX",
    "LAI SUAT": "RATE", "LAISUAT": "RATE", "INTEREST": "RATE",
    "IIP": "IIP", "SAN XUAT CONG NGHIEP": "IIP", "SXCN": "IIP", "CONG NGHIEP": "IIP",
    "BAN LE": "BANLE", "BANLE": "BANLE", "TIEU DUNG": "BANLE", "BAN LE HANG HOA": "BANLE",
    "FDI": "FDI", "DAU TU NUOC NGOAI": "FDI", "VON FDI": "FDI",
    "XNK": "XNK", "XUAT NHAP KHAU": "XNK", "THUONG MAI": "XNK", "CAN CAN THUONG MAI": "XNK",
    "VANG": "GOLD", "GOLD": "GOLD", "GIA VANG": "GOLD",
    "DAU": "OIL", "OIL": "OIL", "GIA DAU": "OIL",
}
MACRO_GROUPS = {
    "CPI":  [("FP.CPI.TOTL.ZG", "Lạm phát CPI (%/năm)", "pct")],
    "GDP":  [("NY.GDP.MKTP.KD.ZG", "Tăng trưởng GDP (%)", "pct"),
             ("NY.GDP.MKTP.CD", "GDP danh nghĩa (tỷ USD)", "usd_bn")],
    "FX":   [("PA.NUS.FCRF", "Tỷ giá chính thức (VND/USD, BQ năm)", "num")],
    "RATE": [("FR.INR.LEND", "Lãi suất cho vay (%/năm)", "pct")],
}


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def macro_key(kw):
    return MACRO_ALIASES.get(strip_accents(kw).upper().strip())


def wb_series(indicator, start=2014, end=None):
    end = end or datetime.today().year
    r = requests.get(WB_BASE.format(ind=indicator),
                     params={"format": "json", "per_page": 200, "date": f"{start}:{end}"},
                     headers={"User-Agent": "Mozilla/5.0"}, timeout=60)
    payload = r.json()
    if not isinstance(payload, list) or len(payload) < 2 or not payload[1]:
        return pd.Series(dtype=float)
    data = {int(x["date"]): x["value"] for x in payload[1] if x["value"] is not None}
    return pd.Series(data).sort_index()


def macro_insights(table, specs):
    bullets = []
    for ind, label, kind in specs:
        s = table[label].dropna()
        if s.empty:
            continue
        last_y, last_v = s.index[-1], s.iloc[-1]
        prev_v = s.iloc[-2] if len(s) > 1 else None
        avg5 = s.tail(5).mean()
        if kind == "pct":
            chg = f"{last_v - prev_v:+.2f} điểm %" if prev_v is not None else "n/a"
            bullets.append(f"{label}: {last_v:.2f}% (năm {last_y}); so năm trước {chg}; BQ 5 năm {avg5:.2f}%.")
        elif kind == "usd_bn":
            chg = f"{pct(last_v, prev_v):+.1f}%" if prev_v else "n/a"
            bullets.append(f"{label}: {last_v:,.0f} tỷ USD (năm {last_y}); so năm trước {chg}.")
        else:
            note = ("VND mất giá" if prev_v and last_v > prev_v else "VND lên giá") if prev_v else ""
            chg = f"{pct(last_v, prev_v):+.1f}%" if prev_v else "n/a"
            bullets.append(f"{label}: {last_v:,.0f} (năm {last_y}); so năm trước {chg} ({note}).")
    return bullets


def build_macro_section(group):
    specs = MACRO_GROUPS[group]
    cols = {}
    for ind, label, kind in specs:
        s = wb_series(ind)
        cols[label] = s / 1e9 if kind == "usd_bn" else s
    table = pd.DataFrame(cols).sort_index()
    if table.dropna(how="all").empty:
        return None
    bullets = macro_insights(table, specs)
    latest = MACRO_LATEST.get(group)
    if latest and latest.get("latest"):
        bullets = [f"MỚI NHẤT ({latest.get('as_of','')}) — {latest['latest']}. "
                   f"Nguồn: {latest.get('source','')}."] + bullets
    out = pd.DataFrame(index=table.index)
    for ind, label, kind in specs:
        out[label] = table[label]
        out[label + (" (Δ điểm %)" if kind == "pct" else " (YoY %)")] = (
            table[label].diff() if kind == "pct" else table[label].pct_change() * 100)
    out = out.tail(10).reset_index().rename(columns={"index": "Năm"})
    codes = ",".join(ind for ind, _, _ in specs)
    return {"key": group, "title": f"VĨ MÔ {group}",
            "source": "World Bank Open Data (Việt Nam) — theo NĂM",
            "source_url": f"https://api.worldbank.org/v2/country/VNM/indicator/{codes}",
            "bullets": bullets, "sheet": (group, [("Theo năm", out, "macro")]), "note": ""}


# ----------------- GHI WORKBOOK -----------------
def _numfmt(cell, header, kind):
    v = cell.value
    if not isinstance(v, (int, float)):
        return
    if header == "Năm":
        cell.number_format = '0'
        return
    if kind in ("stockperiod", "stockraw"):
        if any(k in header for k in ["MoM", "QoQ", "YoY"]):
            cell.value = v / 100.0
            cell.number_format = '+0.0%;-0.0%'
        else:
            cell.number_format = '#,##0'
    elif kind == "compare":
        cell.number_format = '+0.0"%";-0.0"%"'
    else:  # macro
        cell.number_format = '0.00' if any(
            k in header for k in ["%", "Lạm phát", "Lãi suất", "Tăng trưởng"]) else '#,##0'


def _write_block(ws, start_row, title, df, kind):
    """Ghi 1 block (tiêu đề + bảng) từ start_row; trả về dòng trống kế tiếp."""
    from openpyxl.styles import Font, PatternFill
    blue = PatternFill("solid", fgColor="DDEBF7")
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=11, color="1F4E78")
    hr = start_row + 1
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=hr, column=j, value=str(col))
        c.font = Font(bold=True)
        c.fill = blue
    r = hr + 1
    for _, rowdata in df.iterrows():
        for j, col in enumerate(df.columns, start=1):
            v = rowdata[col]
            header = str(col)
            if isinstance(v, float) and pd.isna(v):
                v = None
            elif isinstance(v, pd.Timestamp):
                v = v.to_pydatetime()
            cell = ws.cell(row=r, column=j, value=v)
            if hasattr(v, "year") and not isinstance(v, (int, float)):
                cell.number_format = 'yyyy' if header == "Năm" else 'mm/yyyy'
            else:
                _numfmt(cell, header, kind)
        r += 1
    return r + 1


def _ensure_writable(path):
    """Nếu file đang mở (khóa bởi Excel) -> ghi ra tên khác kèm giờ, khỏi cần đóng Excel."""
    if os.path.exists(path):
        try:
            with open(path, "r+b"):
                pass
        except (PermissionError, OSError):
            base, ext = os.path.splitext(path)
            alt = f"{base}_{datetime.today():%H%M%S}{ext}"
            print(f"  (!) '{os.path.basename(path)}' đang mở (Excel?) → ghi ra '{os.path.basename(alt)}'")
            return alt
    return path


def write_workbook(sections):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    keys = [s["key"] for s in sections]
    fname = (keys[0] if len(keys) == 1 else "TONGHOP") + f"_{datetime.today():%Y%m%d}.xlsx"
    out_path = _ensure_writable(os.path.join(OUT_DIR, fname))

    wb = Workbook()
    ws = wb.active
    ws.title = "Mục lục"
    ws["A1"] = "BÁO CÁO TỔNG HỢP — AUTOMATION 1"
    ws["A1"].font = Font(bold=True, size=15, color="1F4E78")
    ws["A2"] = f"{len(sections)} keyword · Xuất {datetime.today():%d/%m/%Y %H:%M}"
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    NHOM = {"CPI": "Vĩ mô", "GDP": "Vĩ mô", "FX": "Vĩ mô", "RATE": "Vĩ mô",
            "IIP": "Vĩ mô", "BANLE": "Vĩ mô", "FDI": "Vĩ mô", "XNK": "Vĩ mô",
            "GOLD": "Tầng B", "OIL": "Tầng B"}
    hdr = 4
    headers = ["Keyword", "Nhóm", "Nội dung chính", "Nguồn"]
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for j, h in enumerate(headers):
        c = ws.cell(row=hdr, column=1 + j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")
    zebra = PatternFill("solid", fgColor="F2F6FC")
    row = hdr + 1
    for i, sec in enumerate(sections):
        has_sheet = bool(sec.get("sheet"))
        kw_label = sec["key"] + (" ⤵" if has_sheet else "")
        ws.cell(row=row, column=1, value=kw_label).font = Font(bold=True, size=11)
        ws.cell(row=row, column=2, value=NHOM.get(sec["key"], "Cổ phiếu"))
        ws.cell(row=row, column=3, value="\n".join("• " + b for b in sec["bullets"]))
        ws.cell(row=row, column=4, value=sec["source"])
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if i % 2 == 0:
                cell.fill = zebra
        lines = sum(max(1, -(-len(b) // 92)) for b in sec["bullets"]) or 1
        ws.row_dimensions[row].height = max(28, lines * 15 + 4)
        row += 1
    row += 1
    d = ws.cell(row=row, column=1, value="Disclaimer: Báo cáo tự động, tham khảo, không phải khuyến nghị mua/bán. "
               "Chi tiết số liệu xem sheet từng keyword (cột Keyword có ⤵). Nguồn + giờ lấy ở cuối mỗi sheet.")
    d.font = Font(italic=True, size=8, color="808080")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 9
    ws.column_dimensions["C"].width = 92
    ws.column_dimensions["D"].width = 34
    ws.freeze_panes = "A5"

    # Mỗi keyword 1 sheet dữ liệu; các block xếp chồng cho gọn
    stamp = datetime.today().strftime("%d/%m/%Y %H:%M")
    used = set()
    for sec in sections:
        if not sec.get("sheet"):
            continue
        name, blocks = sec["sheet"]
        sname, base, n = name[:31], name[:28], 1
        while sname in used:
            n += 1
            sname = f"{base}_{n}"
        used.add(sname)
        wsd = wb.create_sheet(sname)
        r = 1
        for title, dfb, kind in blocks:
            r = _write_block(wsd, r, title, dfb, kind)
        wsd.column_dimensions["A"].width = 18
        for cl in ["B", "C", "D", "E", "F", "G", "H"]:
            wsd.column_dimensions[cl].width = 15
        src = "Nguồn: " + sec["source"] + (f" — {sec['source_url']}" if sec.get("source_url") else "") \
              + f" · Lấy lúc: {stamp}"
        wsd.cell(row=r + 1, column=1, value=src).font = Font(italic=True, size=8, color="808080")

    wb.save(out_path)
    return out_path


def write_keyword_file(sec):
    """Mỗi keyword 1 file cố định output/<KEY>.xlsx. Sheet 'Nhật ký' CỘNG DỒN mỗi lần chạy
    (1 dòng/ngày, không đè lịch sử); 'Tổng quan' + 'Chi tiết' refresh số mới nhất."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    key = re.sub(r'[\\/:*?\[\]]', "_", sec["key"])
    path = os.path.join(OUT_DIR, f"{key}.xlsx")
    stamp = datetime.today().strftime("%d/%m/%Y %H:%M")
    today = datetime.today().strftime("%d/%m/%Y")

    # Đọc nhật ký cũ (giữ lịch sử), bỏ entry cùng NGÀY (chạy lại trong ngày thì cập nhật)
    old_log = []
    if os.path.exists(path):
        try:
            owb = load_workbook(path)
            if "Nhật ký" in owb.sheetnames:
                for rr in owb["Nhật ký"].iter_rows(min_row=2, values_only=True):
                    if rr and rr[0] and not str(rr[0]).startswith(today):
                        old_log.append((rr[0], rr[1] if len(rr) > 1 else ""))
            owb.close()
        except Exception:
            pass

    wb = Workbook()
    blue = PatternFill("solid", fgColor="DDEBF7")

    # 1) Tổng quan (số mới nhất)
    ws = wb.active
    ws.title = "Tổng quan"
    ws["A1"] = sec["title"]
    ws["A1"].font = Font(bold=True, size=15, color="1F4E78")
    ws["A2"] = f"Cập nhật: {stamp} · Nguồn: {sec['source']}"
    ws["A2"].font = Font(italic=True, size=9, color="808080")
    r = 4
    for b in sec["bullets"]:
        c = ws.cell(row=r, column=1, value="• " + b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = max(15, (-(-len(b) // 95)) * 15)
        r += 1
    r += 1
    if sec.get("source_url"):
        ws.cell(row=r, column=1, value="Nguồn: " + sec["source"] + " — " + sec["source_url"]).font = Font(italic=True, size=8, color="808080")
        r += 1
    ws.cell(row=r, column=1, value="Disclaimer: Báo cáo tự động, tham khảo, không phải khuyến nghị mua/bán.").font = Font(italic=True, size=8, color="808080")
    ws.column_dimensions["A"].width = 100

    # 2) Nhật ký (cộng dồn)
    log = wb.create_sheet("Nhật ký")
    for j, h in enumerate(["Ngày chạy", "Nội dung đã ghi"], start=1):
        c = log.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = blue
    content = "\n".join("• " + b for b in sec["bullets"])
    rr = 2
    for (d, ct) in old_log + [(stamp, content)]:
        log.cell(row=rr, column=1, value=d)
        cc = log.cell(row=rr, column=2, value=ct)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        log.row_dimensions[rr].height = max(15, (str(ct).count("\n") + 1) * 15)
        rr += 1
    log.column_dimensions["A"].width = 16
    log.column_dimensions["B"].width = 100
    log.freeze_panes = "A2"

    # 3) Chi tiết (block xếp chồng, số mới nhất)
    if sec.get("sheet"):
        name, blocks = sec["sheet"]
        det = wb.create_sheet("Chi tiết")
        rb = 1
        for title, dfb, kind in blocks:
            rb = _write_block(det, rb, title, dfb, kind)
        det.column_dimensions["A"].width = 18
        for cl in ["B", "C", "D", "E", "F", "G", "H"]:
            det.column_dimensions[cl].width = 15
        det.cell(row=rb + 1, column=1, value="Nguồn: " + sec["source"]
                 + (f" — {sec['source_url']}" if sec.get("source_url") else "") + f" · Lấy lúc: {stamp}"
                 ).font = Font(italic=True, size=8, color="808080")

    try:
        wb.save(path)
    except PermissionError:
        base, ext = os.path.splitext(path)
        path = f"{base}_{datetime.today():%H%M%S}{ext}"
        wb.save(path)
        print(f"    (!) '{key}.xlsx' đang mở → ghi tạm '{os.path.basename(path)}' (nhật ký chưa gộp được).")
    return path


# ----------------- MAIN -----------------
def main():
    global FULL
    args = [a.strip() for a in sys.argv[1:] if a.strip()]
    if "--full" in args or "-f" in args:
        FULL = True
        args = [a for a in args if a not in ("--full", "-f")]
    raw = args or ["HPG"]
    seen, kws = set(), []
    for k in raw:  # khử trùng lặp, giữ thứ tự
        if k.lower() not in seen:
            seen.add(k.lower()); kws.append(k)
    print(f"[Automation 1] {len(kws)} keyword: {', '.join(kws)}")

    # Prefetch vốn hóa cho mọi mã cổ phiếu trong 1 request
    stock_tks = [k.upper() for k in kws if not macro_key(k) and 2 <= len(k) <= 4 and k.isalpha()]
    if stock_tks:
        prefetch_listed_shares(stock_tks)

    sections = []
    for i, kw in enumerate(kws):
        if i:
            time.sleep(0.8)  # giãn nhịp tránh rate-limit nguồn
        grp = macro_key(kw)
        if grp in ("GOLD", "OIL"):
            ten = "vàng" if grp == "GOLD" else "dầu"
            print(f"  • {kw}: giá {ten} bị chặn từ Python local → dùng luồng Claude-fetch "
                  f"(snapshot_report.py). Bỏ qua trong workbook này.")
            sections.append({"key": grp, "title": f"GIÁ {ten.upper()}", "source": "Cần Claude fetch",
                             "bullets": [f"Giá {ten} không lấy được từ Python local (nguồn chặn 403/JS). "
                                         "Dùng: Claude fetch → data/<KEY>.json → snapshot_report.py."],
                             "sheet": None, "note": ""})
            continue
        try:
            if grp and grp in MACRO_GROUPS:
                print(f"  • {kw} → vĩ mô {grp} (World Bank + số mới nhất)")
                sec = build_macro_section(grp)
                if sec is None:
                    print(f"    (!) Không lấy được dữ liệu World Bank cho {grp}.")
                    continue
            elif grp and grp in MACRO_LATEST:
                print(f"  • {kw} → vĩ mô {grp} (số mới nhất, nso)")
                sec = build_latest_only_section(grp)
                if sec is None:
                    print(f"    (!) Chưa có số mới nhất cho {grp}.")
                    continue
            else:
                tk = kw.upper()
                if not (2 <= len(tk) <= 4 and tk.isalpha()):
                    print(f"    (!) '{kw}' không nhận diện được (mã CP hoặc vĩ mô). Bỏ qua.")
                    continue
                print(f"  • {kw} → cổ phiếu (vnstock)")
                sec = build_stock_section(tk)
            sections.append(sec)
        except Exception as e:
            print(f"    (!) Lỗi xử lý '{kw}': {type(e).__name__}: {e}")

    sections = [s for s in sections if s]
    if not sections:
        print("Không có keyword hợp lệ. Ví dụ: kw CPI GDP HPG")
        return

    print("\n  --- INSIGHTS ---")
    for sec in sections:
        print(f"  [{sec['title']}]")
        for b in sec["bullets"]:
            print("   • " + b)
    print("\n  Đã xuất (mỗi keyword 1 file, nhật ký cộng dồn):")
    for sec in sections:
        print("   - " + write_keyword_file(sec))


if __name__ == "__main__":
    main()
